from datetime import datetime
from pathlib import Path

FIXED_INVOICE_PATH = Path("data/hoa_don_robot_demo.xlsx")


def _ensure_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        return Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter
    except ImportError as exc:
        raise ImportError("Thiếu openpyxl. Cài bằng: pip install openpyxl") from exc


def _style_header(ws, row=1, max_col=8):
    Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter = _ensure_openpyxl()
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_table_style(ws, min_row, max_row, min_col, max_col):
    Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter = _ensure_openpyxl()
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autosize(ws, max_width=42):
    Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter = _ensure_openpyxl()
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def _create_workbook(path: Path):
    Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter = _ensure_openpyxl()
    wb = Workbook()

    ws = wb.active
    ws.title = "HoaDonHienTai"
    ws.merge_cells("A1:H1")
    ws["A1"] = "HÓA ĐƠN HIỆN TẠI - ROBOT PHỤC VỤ NHÀ HÀNG"
    ws["A1"].font = Font(size=16, bold=True, color="102A43")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["File cố định", str(path), "", "", "", "", "", ""])
    ws.append(["Trạng thái", "EMPTY", "", "", "", "", "", ""])
    ws.append(["Cập nhật lần cuối", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", "", "", ""])
    ws.append([])
    ws.append(["STT", "Mã món", "Tên món", "Đơn giá", "Số lượng", "Thành tiền", "Trạng thái", "Ghi chú"])
    _style_header(ws, 7, 8)

    ws2 = wb.create_sheet("LichSuThanhToan")
    ws2.append(["Thời gian", "Mã thanh toán", "Trạng thái", "Tổng tiền", "Số món", "QR/URL", "Ghi chú", "Chi tiết"])
    _style_header(ws2, 1, 8)

    ws3 = wb.create_sheet("ChiTietDaThanhToan")
    ws3.append(["Thời gian", "Mã thanh toán", "STT", "Mã món", "Tên món", "Đơn giá", "Số lượng", "Thành tiền", "Trạng thái"])
    _style_header(ws3, 1, 9)

    ws4 = wb.create_sheet("EventLog")
    ws4.append(["Thời gian", "Nguồn", "Action", "Intent", "Nội dung"])
    _style_header(ws4, 1, 5)

    for sheet in wb.worksheets:
        _autosize(sheet)
        sheet.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def sync_invoice_excel(
    order_manager,
    event_logs,
    payment_status="UNPAID",
    payment_id="",
    payment_url="",
    note="",
    append_payment=False,
    output_path=None
):
    """
    Ghi/đồng bộ hóa đơn vào file Excel cố định.

    - Khi bấm Xác nhận đơn: ghi sheet HoaDonHienTai với trạng thái CONFIRMED_WAITING_PAYMENT.
    - Khi QR xác nhận đã thanh toán: append vào LichSuThanhToan và ChiTietDaThanhToan.
    - Sau đó app reset hóa đơn, nhưng lịch sử thanh toán vẫn được giữ trong file Excel.
    """
    Workbook, load_workbook, Font, PatternFill, Border, Side, Alignment, get_column_letter = _ensure_openpyxl()

    path = Path(output_path) if output_path else FIXED_INVOICE_PATH
    if not path.exists():
        _create_workbook(path)

    wb = load_workbook(path)

    # Đảm bảo đủ sheet.
    for name in ["HoaDonHienTai", "LichSuThanhToan", "ChiTietDaThanhToan", "EventLog"]:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    orders = order_manager.as_list()
    total = order_manager.total()

    # Sheet hóa đơn hiện tại: xóa nội dung cũ và ghi lại.
    ws = wb["HoaDonHienTai"]
    ws.delete_rows(1, ws.max_row)

    ws.merge_cells("A1:H1")
    ws["A1"] = "HÓA ĐƠN HIỆN TẠI - ROBOT PHỤC VỤ NHÀ HÀNG"
    ws["A1"].font = Font(size=16, bold=True, color="102A43")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["File cố định", str(path), "", "", "", "", "", ""])
    ws.append(["Trạng thái", payment_status, "", "", "", "", "", ""])
    ws.append(["Mã thanh toán", payment_id, "", "", "", "", "", ""])
    ws.append(["QR/URL", payment_url, "", "", "", "", "", ""])
    ws.append(["Cập nhật lần cuối", now, "", "", "", "", "", ""])
    ws.append([])
    ws.append(["STT", "Mã món", "Tên món", "Đơn giá", "Số lượng", "Thành tiền", "Trạng thái", "Ghi chú"])
    header_row = 9
    _style_header(ws, header_row, 8)

    for idx, item in enumerate(orders, start=1):
        ws.append([
            idx,
            item.get("item_id", ""),
            item.get("name", ""),
            item.get("price", 0),
            item.get("quantity", 0),
            item.get("subtotal", 0),
            "Đã xác nhận" if item.get("status") == "confirmed" else "Chờ xác nhận",
            note
        ])

    summary_row = header_row + len(orders) + 2
    ws.cell(summary_row, 5, "Tổng tiền:").font = Font(bold=True)
    ws.cell(summary_row, 6, total).font = Font(bold=True)
    ws.cell(summary_row, 6).number_format = '#,##0" đ"'

    ws.cell(summary_row + 1, 5, "Đã xác nhận:")
    ws.cell(summary_row + 1, 6, order_manager.confirmed_total()).number_format = '#,##0" đ"'
    ws.cell(summary_row + 2, 5, "Chờ xác nhận:")
    ws.cell(summary_row + 2, 6, order_manager.pending_total()).number_format = '#,##0" đ"'

    if orders:
        _apply_table_style(ws, header_row, header_row + len(orders), 1, 8)
    _apply_table_style(ws, summary_row, summary_row + 2, 5, 6)

    for row in range(header_row + 1, header_row + len(orders) + 1):
        ws.cell(row, 4).number_format = '#,##0" đ"'
        ws.cell(row, 6).number_format = '#,##0" đ"'

    ws.freeze_panes = "A10"

    # Append lịch sử nếu thanh toán thành công.
    if append_payment:
        history = wb["LichSuThanhToan"]
        if history.max_row < 1:
            history.append(["Thời gian", "Mã thanh toán", "Trạng thái", "Tổng tiền", "Số món", "QR/URL", "Ghi chú", "Chi tiết"])
            _style_header(history, 1, 8)

        detail_text = "; ".join([f'{i.get("name")} x{i.get("quantity")}' for i in orders])
        history.append([now, payment_id, payment_status, total, len(orders), payment_url, note, detail_text])
        last = history.max_row
        history.cell(last, 4).number_format = '#,##0" đ"'
        _apply_table_style(history, last, last, 1, 8)

        detail = wb["ChiTietDaThanhToan"]
        if detail.max_row < 1:
            detail.append(["Thời gian", "Mã thanh toán", "STT", "Mã món", "Tên món", "Đơn giá", "Số lượng", "Thành tiền", "Trạng thái"])
            _style_header(detail, 1, 9)

        for idx, item in enumerate(orders, start=1):
            detail.append([
                now,
                payment_id,
                idx,
                item.get("item_id", ""),
                item.get("name", ""),
                item.get("price", 0),
                item.get("quantity", 0),
                item.get("subtotal", 0),
                "ĐÃ THANH TOÁN"
            ])
            row = detail.max_row
            detail.cell(row, 6).number_format = '#,##0" đ"'
            detail.cell(row, 8).number_format = '#,##0" đ"'
            _apply_table_style(detail, row, row, 1, 9)

    # Ghi event log mới nhất.
    log_ws = wb["EventLog"]
    if log_ws.max_row < 1:
        log_ws.append(["Thời gian", "Nguồn", "Action", "Intent", "Nội dung"])
        _style_header(log_ws, 1, 5)

    # Ghi tối đa 10 log mới nhất mỗi lần để tiện theo dõi.
    for log in event_logs[-10:]:
        log_ws.append([
            log.get("time", now),
            log.get("source", ""),
            log.get("action", ""),
            log.get("intent", ""),
            log.get("reply", "")
        ])
        _apply_table_style(log_ws, log_ws.max_row, log_ws.max_row, 1, 5)

    for sheet in wb.worksheets:
        _autosize(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return str(path)


def reset_current_invoice_sheet(event_logs=None, output_path=None):
    """
    Sau khi thanh toán xong và reset bill trên GUI, sheet HoaDonHienTai sẽ về trạng thái EMPTY.
    Lịch sử thanh toán vẫn giữ nguyên.
    """
    class EmptyOrderManager:
        def as_list(self): return []
        def total(self): return 0
        def confirmed_total(self): return 0
        def pending_total(self): return 0

    return sync_invoice_excel(
        EmptyOrderManager(),
        event_logs or [],
        payment_status="EMPTY",
        note="Hóa đơn đã reset sau thanh toán.",
        output_path=output_path
    )
