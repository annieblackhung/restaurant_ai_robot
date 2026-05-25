from datetime import datetime
from pathlib import Path

FIXED_EXCEL_PATH = Path("data/hoa_don_robot_demo.xlsx")

INVOICE_HEADERS = [
    "Mã hóa đơn",
    "Thời gian xác nhận",
    "Mã món",
    "Tên món",
    "Đơn giá",
    "Số lượng",
    "Thành tiền",
    "Trạng thái đơn",
    "Thanh toán",
    "Thời gian thanh toán",
]

EVENT_HEADERS = [
    "Thời gian",
    "Mã hóa đơn",
    "Sự kiện",
    "Nội dung",
]


def _load_or_create_workbook():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("Thiếu openpyxl. Cài bằng: pip install openpyxl") from exc

    FIXED_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if FIXED_EXCEL_PATH.exists():
        wb = load_workbook(FIXED_EXCEL_PATH)
    else:
        wb = Workbook()

    if "HoaDon" not in wb.sheetnames:
        ws = wb.active
        ws.title = "HoaDon"
    else:
        ws = wb["HoaDon"]

    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(INVOICE_HEADERS)

    # Nếu file cũ thiếu cột "Thanh toán", thêm cột vào cuối nhưng vẫn giữ dữ liệu cũ.
    existing_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for header in INVOICE_HEADERS:
        if header not in existing_headers:
            ws.cell(1, ws.max_column + 1, header)
            existing_headers.append(header)

    if "SuKien" not in wb.sheetnames:
        log_ws = wb.create_sheet("SuKien")
        log_ws.append(EVENT_HEADERS)
    else:
        log_ws = wb["SuKien"]
        if log_ws.max_row == 1 and log_ws.cell(1, 1).value is None:
            log_ws.append(EVENT_HEADERS)

    _format_workbook(wb)
    return wb


def _format_workbook(wb):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "HoaDon": {
            "A": 18, "B": 22, "C": 14, "D": 22, "E": 14,
            "F": 12, "G": 16, "H": 18, "I": 18, "J": 22
        },
        "SuKien": {
            "A": 22, "B": 18, "C": 22, "D": 60
        }
    }

    for sheet_name, config in widths.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col, width in config.items():
                ws.column_dimensions[col].width = width

    if "HoaDon" in wb.sheetnames:
        ws = wb["HoaDon"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 5).number_format = '#,##0" đ"'
            ws.cell(row, 7).number_format = '#,##0" đ"'


def _header_map(ws):
    return {
        ws.cell(1, col).value: col
        for col in range(1, ws.max_column + 1)
    }


def append_invoice(order_id: str, order_items: list) -> str:
    """
    Ghi đơn mới vào Excel.
    Trạng thái thanh toán ban đầu luôn là 'Chưa thanh toán'.
    Không xóa lịch sử hóa đơn cũ.
    """
    if not order_items:
        raise ValueError("Không có món nào để ghi hóa đơn.")

    wb = _load_or_create_workbook()
    ws = wb["HoaDon"]
    headers = _header_map(ws)

    # Tránh ghi trùng nếu người dùng bấm xác nhận nhiều lần cho cùng mã hóa đơn.
    id_col = headers["Mã hóa đơn"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, id_col).value == order_id:
            wb.save(FIXED_EXCEL_PATH)
            return str(FIXED_EXCEL_PATH)

    confirmed_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for item in order_items:
        row_data = {
            "Mã hóa đơn": order_id,
            "Thời gian xác nhận": confirmed_time,
            "Mã món": item.get("item_id", ""),
            "Tên món": item.get("name", ""),
            "Đơn giá": item.get("price", 0),
            "Số lượng": item.get("quantity", 0),
            "Thành tiền": item.get("subtotal", 0),
            "Trạng thái đơn": "Đã xác nhận",
            "Thanh toán": "Chưa thanh toán",
            "Thời gian thanh toán": "",
        }

        next_row = ws.max_row + 1
        for header, value in row_data.items():
            ws.cell(next_row, headers[header], value)

    append_event(order_id, "XÁC NHẬN ĐƠN", f"Ghi hóa đơn {order_id} vào Excel", wb=wb)
    _format_workbook(wb)
    wb.save(FIXED_EXCEL_PATH)
    return str(FIXED_EXCEL_PATH)


def update_payment_status(order_id: str, paid: bool = True) -> str:
    """
    Cập nhật cột 'Thanh toán' cho hóa đơn đã có.
    Không xóa bất kỳ hóa đơn cũ nào.
    """
    wb = _load_or_create_workbook()
    ws = wb["HoaDon"]
    headers = _header_map(ws)

    payment_status = "Đã thanh toán" if paid else "Chưa thanh toán"
    paid_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if paid else ""

    id_col = headers["Mã hóa đơn"]
    payment_col = headers["Thanh toán"]
    paid_time_col = headers["Thời gian thanh toán"]

    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, id_col).value == order_id:
            ws.cell(row, payment_col, payment_status)
            ws.cell(row, paid_time_col, paid_time)
            found = True

    if not found:
        raise ValueError(f"Không tìm thấy mã hóa đơn trong Excel: {order_id}")

    append_event(order_id, "THANH TOÁN", f"Hóa đơn {order_id}: {payment_status}", wb=wb)
    _format_workbook(wb)
    wb.save(FIXED_EXCEL_PATH)
    return str(FIXED_EXCEL_PATH)


def append_event(order_id: str, event_name: str, content: str, wb=None) -> str:
    owns_workbook = wb is None
    if wb is None:
        wb = _load_or_create_workbook()

    log_ws = wb["SuKien"]
    log_ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        order_id,
        event_name,
        content,
    ])

    if owns_workbook:
        _format_workbook(wb)
        wb.save(FIXED_EXCEL_PATH)

    return str(FIXED_EXCEL_PATH)


def get_fixed_excel_path() -> str:
    wb = _load_or_create_workbook()
    _format_workbook(wb)
    wb.save(FIXED_EXCEL_PATH)
    return str(FIXED_EXCEL_PATH)
